"""管理端 endpoint：统计仪表盘数据 + 导出对话日志为 Excel。

鉴权：设置了 ADMIN_TOKEN 时，所有 /admin/* 需带 X-Admin-Token 请求头
或 ?token= 参数；未设置时不校验（本地开发模式）。公网部署务必设置。
"""
import hmac
from io import BytesIO

from fastapi import APIRouter, Depends, Header, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..config import get_settings
from ..db import get_conn


def require_admin(
    x_admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
    token: str | None = Query(default=None),
) -> None:
    s = get_settings()
    if not s.admin_token:
        return  # 未配置 → 开发模式不校验
    supplied = x_admin_token or token or ""
    if not hmac.compare_digest(supplied, s.admin_token):
        raise HTTPException(status_code=401, detail="admin token 无效")


router = APIRouter(prefix="/admin", tags=["admin"], dependencies=[Depends(require_admin)])


_HEADERS = ["id", "时间", "会话", "角色", "意图", "风险", "日志级别", "内容"]
_COL_WIDTHS = [6, 20, 14, 10, 12, 8, 12, 80]

_LEVEL_FILL = {
    "ALERT": PatternFill("solid", fgColor="FFC7CE"),
    "WARN": PatternFill("solid", fgColor="FFEB9C"),
    "INFO": PatternFill("solid", fgColor="FFFFFF"),
}


@router.get("/export_excel")
def export_excel(
    session_id: str | None = Query(default=None, description="只导某个 session"),
    log_level: str | None = Query(default=None, description="过滤 INFO/WARN/ALERT"),
    limit: int = Query(default=1000, ge=1, le=10000),
):
    sql = "SELECT id, created_at, session_id, role, intent, risk_level, log_level, content FROM messages"
    conds: list[str] = []
    args: list = []
    if session_id:
        conds.append("session_id = ?")
        args.append(session_id)
    if log_level:
        conds.append("log_level = ?")
        args.append(log_level.upper())
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT ?"
    args.append(limit)

    with get_conn() as conn:
        rows = conn.execute(sql, args).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "对话日志"

    # 表头
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="305496")

    # 列宽
    for i, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 数据行
    for r in rows:
        ws.append(
            [
                r["id"],
                r["created_at"],
                r["session_id"],
                r["role"],
                r["intent"] or "",
                r["risk_level"] or "",
                r["log_level"] or "",
                r["content"],
            ]
        )
        # 按日志级别给整行上色
        level = (r["log_level"] or "").upper()
        fill = _LEVEL_FILL.get(level)
        if fill is not None and level != "INFO":
            row_idx = ws.max_row
            for col in range(1, len(_HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # 内容列自动换行
    for row in ws.iter_rows(min_row=2):
        row[-1].alignment = Alignment(wrap_text=True, vertical="top")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "eldercare_log.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/stats")
def stats():
    """仪表盘数据：总量、意图/级别分布、14 天趋势、最近告警。"""
    with get_conn() as conn:
        total = conn.execute("SELECT COUNT(*) AS c FROM messages").fetchone()["c"]
        sessions = conn.execute(
            "SELECT COUNT(DISTINCT session_id) AS c FROM messages"
        ).fetchone()["c"]
        users = conn.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
        alerts = conn.execute(
            "SELECT COUNT(*) AS c FROM messages WHERE role='user' AND log_level='ALERT'"
        ).fetchone()["c"]
        by_intent = conn.execute(
            "SELECT intent, COUNT(*) AS c FROM messages "
            "WHERE role='user' AND intent IS NOT NULL GROUP BY intent"
        ).fetchall()
        by_level = conn.execute(
            "SELECT log_level, COUNT(*) AS c FROM messages "
            "WHERE role='user' AND log_level IS NOT NULL GROUP BY log_level"
        ).fetchall()
        by_day = conn.execute(
            "SELECT date(created_at) AS d, COUNT(*) AS c FROM messages "
            "WHERE role='user' AND created_at >= datetime('now', '-13 days') "
            "GROUP BY date(created_at) ORDER BY d"
        ).fetchall()
        recent_alerts = conn.execute(
            "SELECT created_at, session_id, intent, risk_level, log_level, "
            "       substr(content, 1, 60) AS content "
            "FROM messages WHERE role='user' AND log_level IN ('WARN','ALERT') "
            "ORDER BY id DESC LIMIT 10"
        ).fetchall()

    return {
        "total_messages": total,
        "total_sessions": sessions,
        "total_users": users,
        "total_alerts": alerts,
        "by_intent": {r["intent"]: r["c"] for r in by_intent},
        "by_log_level": {r["log_level"]: r["c"] for r in by_level},
        "by_day": [{"date": r["d"], "count": r["c"]} for r in by_day],
        "recent_alerts": [dict(r) for r in recent_alerts],
    }
